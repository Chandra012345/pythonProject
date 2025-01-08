import openpyxl
from openpyxl.styles import PatternFill

def getRowCount(file,sheetName):
    workbook= openpyxl.load_workbook(file)
    sheet= workbook[sheetName]
    return (sheet.max_row)


def getColumnCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return (sheet.max_column)

def readData(file,sheetName,rownum,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return sheet.cell(rownum,columnno).value

def writeData(file,sheetName,rownum,columnno,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    sheet.cell(rownum,columnno).value=data
    workbook.save(file)

def fillGreenColor(file,sheetName,rownum,coulmnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    greenFill=PatternFill(start_color='60b212',end_color='60b212',fill_type='solid')
    sheet.cell(rownum,coulmnno).fill=greenFill
    workbook.save(file)



def fillRedColor(file, sheet_name, row_num, column_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    redFill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
    sheet.cell(row=row_num, column=column_num).fill = redFill
    workbook.save(file)



