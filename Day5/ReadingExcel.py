import openpyxl
#file-- workbook--sheets--rows--cells
file="C:\\Users\\seleniumpractice.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]

rows=sheet.max_row   #counts the number of rows
cols=sheet.max_column  # counts the number of rows

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value, end=" ")
    print()


